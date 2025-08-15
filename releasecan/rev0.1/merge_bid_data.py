# merge_bid_data.py

import pandas as pd
import logging
from openpyxl import load_workbook

logging.basicConfig(filename="merge_log.txt", level=logging.INFO)

def merge_preserving_format(existing_path, new_data_path, output_path=None):
    new_df = pd.read_excel(new_data_path)
    new_df["施工番号"] = new_df["施工番号"].astype(str)

    existing_df = pd.read_excel(existing_path)  
    wb = load_workbook(existing_path)
    ws = wb["Sheet1"] # または wb["Sheet名"]

    # 施工番号列のインデックスを取得
    header = [cell.value for cell in ws[1]]
    seko_col = header.index("施工番号・工事番号") + 1

    max_row = ws.max_row
    existing_seko_nos = {
        str(ws.cell(row=i, column=seko_col).value): i 
        for i in range(2, max_row + 1)
    }

    for _, row in new_df.iterrows():
        seko_no = row["施工番号"]
        if seko_no in existing_seko_nos:
            r = existing_seko_nos[seko_no]
            ws.cell(row=r, column=header.index("工事名") + 1).value = row["工事・業務名"]
            ws.cell(row=r, column=header.index("入札年月日") + 1).value =  row["開札予定日"]
            ws.cell(row=r, column=header.index("予定価格(税込)") + 1).value = row["予定価格"]
            ws.cell(row=r, column=header.index("落札価格") + 1).value = row["落札金額"]
            ws.cell(row=r, column=header.index("業者コード") + 1).value = row["落札業者"]
            ws.cell(row=r, column=header.index("工事場所") + 1).value = row["工事場所"]
            logging.info(f"更新: {seko_no}")
        else:
            new_row = [None] * len(header)
            new_row[seko_col - 1] = seko_no
            new_row[header.index("工事名")] = row["工事・業務名"]
            new_row[header.index("入札年月日")] = row["開札予定日"]
            new_row[header.index("予定価格(税込)")] = row["予定価格"]
            new_row[header.index("落札価格")] = row["落札金額"]
            new_row[header.index("業者コード")] = row["落札業者"]
            new_row[header.index("工事場所")] = row["工事場所"]
            ws.append(new_row)
            logging.info(f"追加: {seko_no}")
            
            max_number = pd.to_numeric(existing_df["仕分番号"], errors="coerce").max()
            new_row_df = pd.DataFrame([dict(zip(header, new_row))])
            new_row_df["仕分番号"] = int(max_number) + 1 if pd.notnull(max_number) else 1
            existing_df = pd.concat([existing_df, new_row_df], ignore_index=True)
        
        wb.save(existing_path)        
        print(f"✅ 書式を保持したまま統合完了: {existing_path}")
        
    output_path = output_path or existing_path
    existing_df.to_excel(output_path, index=False)
    print(f"✅ 統合完了: {output_path}")
